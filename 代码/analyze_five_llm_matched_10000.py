# -*- coding: utf-8 -*-
from __future__ import annotations

import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import font_manager
from openpyxl import load_workbook

from exclude_ai_error_from_raw import ROOT, ensure_unique_dir


GROUP_ORDER = ["Total", "X-ray", "CT", "MRI"]
LABEL_ORDER = ["a", "b", "c", "d", "e", "f"]
ERROR_LABELS = {"a", "b", "c", "d", "e"}
METRIC_COLUMNS = ["ACC", "Sensitivity", "Precision", "Specificity", "F1 score"]
MODEL_ORDER = ["Qwen 3", "GPT-5.4 Pro", "Gemini 3.1 Pro", "Llama 4", "DeepSeek-V3.1"]
PLOT_METRIC_COLORS = ["#163A5F", "#356CA5", "#6797C8", "#9FC0E0", "#D88C5A"]
GROUP_COLORS = {"Total": "#163A5F", "X-ray": "#356CA5", "CT": "#6797C8", "MRI": "#D88C5A"}
ROOT_MODEL_SPECS = {
    "Qwen 3": {"slot_pairs": [(6, 7), (11, 12), (16, 17)]},
    "Llama 4": {"slot_pairs": [(6, 8), (11, 13), (16, 18)]},
    "DeepSeek-V3.1": {"slot_pairs": [(6, 9), (11, 14), (16, 19)]},
}
SUBDIR_MODEL_SPECS = {
    "j_gpt-5.4": {"model_name": "GPT-5.4 Pro", "slot_pairs": [(6, 7), (9, 10), (12, 13)]},
    "j_gemini-3.1-pro-high": {"model_name": "Gemini 3.1 Pro", "slot_pairs": [(6, 7), (9, 10), (12, 13)]},
}
MATCHED_LIMITS = {"CT": 4000, "MRI": 3000, "X-ray": 3000}


@dataclass
class ReportRecord:
    model_name: str
    modality: str
    report_key: str
    pred_has_error: bool
    true_has_error: bool


@dataclass
class InstanceRecord:
    model_name: str
    modality: str
    report_key: str
    slot: int
    y_true: str
    y_pred: str


def configure_plot_style() -> None:
    available_fonts = {font.name for font in font_manager.fontManager.ttflist}
    for preferred in ["Arial", "Times New Roman", "Calibri", "DejaVu Sans"]:
        if preferred in available_fonts:
            plt.rcParams["font.family"] = preferred
            break
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["font.size"] = 10.5
    plt.rcParams["axes.titlesize"] = 14
    plt.rcParams["axes.labelsize"] = 11
    plt.rcParams["legend.fontsize"] = 9.5
    plt.rcParams["xtick.labelsize"] = 9.5
    plt.rcParams["ytick.labelsize"] = 9.5


def normalize_label(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text or text == "nan":
        return ""
    if text in LABEL_ORDER:
        return text
    if text in {"n", "none", "no error"}:
        return "f"
    return text


def safe_divide(numerator: float, denominator: float) -> float:
    if denominator == 0:
        return float("nan")
    return numerator / denominator


def format_metric(value: float) -> float | None:
    if isinstance(value, float) and math.isnan(value):
        return None
    return round(float(value), 6)


def detect_ai_root() -> Path:
    llm_root = next(
        p for p in ROOT.iterdir()
        if p.is_dir() and p.name.startswith("LLM") and any(c.is_dir() and c.name.startswith("AI") for c in p.iterdir())
    )
    return next(c for c in llm_root.iterdir() if c.is_dir() and c.name.startswith("AI"))


def modality_from_name(file_name: str) -> str:
    upper_name = file_name.upper()
    if "X-RAY" in upper_name or "DR" in upper_name:
        return "X-ray"
    if "CT" in upper_name:
        return "CT"
    if "MRI" in upper_name:
        return "MRI"
    raise ValueError(f"Cannot detect modality from file name: {file_name}")


def iter_rows(path: Path, limit: int | None = None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    row_iter = ws.iter_rows(values_only=True)
    next(row_iter, None)
    count = 0
    for row in row_iter:
        yield row
        count += 1
        if limit is not None and count >= limit:
            break
    workbook.close()


def extract_records_from_rows(rows_iter, model_name: str, modality: str, report_prefix: str, slot_pairs: list[tuple[int, int]]):
    report_records: list[ReportRecord] = []
    instance_records: list[InstanceRecord] = []
    true_counter = Counter()
    file_report_rows = 0
    file_instance_rows = 0

    for row_number, row in enumerate(rows_iter, start=2):
        report_key = f"{report_prefix}::{row_number}::{model_name}"
        pred_labels: list[str] = []
        true_labels: list[str] = []
        slot_count = 0

        for slot_no, (true_idx, pred_idx) in enumerate(slot_pairs, start=1):
            y_true = normalize_label(row[true_idx] if true_idx < len(row) else None)
            y_pred = normalize_label(row[pred_idx] if pred_idx < len(row) else None)
            if not y_true and not y_pred:
                continue
            y_true = y_true or "f"
            y_pred = y_pred or "f"
            if y_true not in LABEL_ORDER or y_pred not in LABEL_ORDER:
                continue

            instance_records.append(
                InstanceRecord(
                    model_name=model_name,
                    modality=modality,
                    report_key=report_key,
                    slot=slot_no,
                    y_true=y_true,
                    y_pred=y_pred,
                )
            )
            slot_count += 1
            true_labels.append(y_true)
            pred_labels.append(y_pred)
            true_counter[y_true] += 1

        if slot_count == 0:
            instance_records.append(
                InstanceRecord(
                    model_name=model_name,
                    modality=modality,
                    report_key=report_key,
                    slot=1,
                    y_true="f",
                    y_pred="f",
                )
            )
            slot_count = 1
            true_labels.append("f")
            pred_labels.append("f")
            true_counter["f"] += 1

        report_records.append(
            ReportRecord(
                model_name=model_name,
                modality=modality,
                report_key=report_key,
                pred_has_error=any(label in ERROR_LABELS for label in pred_labels),
                true_has_error=any(label in ERROR_LABELS for label in true_labels),
            )
        )
        file_report_rows += 1
        file_instance_rows += slot_count

    return report_records, instance_records, true_counter, file_report_rows, file_instance_rows


def load_matched_records() -> tuple[list[ReportRecord], list[InstanceRecord], pd.DataFrame]:
    source_root = detect_ai_root()
    report_records: list[ReportRecord] = []
    instance_records: list[InstanceRecord] = []
    file_stats_rows: list[dict[str, object]] = []

    root_files = sorted(p for p in source_root.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$"))
    for path in root_files:
        modality = modality_from_name(path.name)
        limit = MATCHED_LIMITS[modality]
        for model_name, spec in ROOT_MODEL_SPECS.items():
            r_records, i_records, true_counter, file_report_rows, file_instance_rows = extract_records_from_rows(
                iter_rows(path, limit=limit),
                model_name=model_name,
                modality=modality,
                report_prefix=path.name,
                slot_pairs=spec["slot_pairs"],
            )
            report_records.extend(r_records)
            instance_records.extend(i_records)
            file_stats_rows.append(
                {
                    "Model": model_name,
                    "SourceType": "root_matched_subset",
                    "File": path.name,
                    "Modality": modality,
                    "MatchedReportRows": file_report_rows,
                    "InstanceRows": file_instance_rows,
                    "TrueLabel_a": true_counter.get("a", 0),
                    "TrueLabel_b": true_counter.get("b", 0),
                    "TrueLabel_c": true_counter.get("c", 0),
                    "TrueLabel_d": true_counter.get("d", 0),
                    "TrueLabel_e": true_counter.get("e", 0),
                    "TrueLabel_f": true_counter.get("f", 0),
                }
            )

    for folder_name, spec in SUBDIR_MODEL_SPECS.items():
        model_dir = source_root / folder_name
        for path in sorted(p for p in model_dir.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")):
            modality = modality_from_name(path.name)
            r_records, i_records, true_counter, file_report_rows, file_instance_rows = extract_records_from_rows(
                iter_rows(path, limit=None),
                model_name=spec["model_name"],
                modality=modality,
                report_prefix=path.name,
                slot_pairs=spec["slot_pairs"],
            )
            report_records.extend(r_records)
            instance_records.extend(i_records)
            file_stats_rows.append(
                {
                    "Model": spec["model_name"],
                    "SourceType": folder_name,
                    "File": path.name,
                    "Modality": modality,
                    "MatchedReportRows": file_report_rows,
                    "InstanceRows": file_instance_rows,
                    "TrueLabel_a": true_counter.get("a", 0),
                    "TrueLabel_b": true_counter.get("b", 0),
                    "TrueLabel_c": true_counter.get("c", 0),
                    "TrueLabel_d": true_counter.get("d", 0),
                    "TrueLabel_e": true_counter.get("e", 0),
                    "TrueLabel_f": true_counter.get("f", 0),
                }
            )

    return report_records, instance_records, pd.DataFrame(file_stats_rows)


def summarize_report_binary(report_records: list[ReportRecord]) -> pd.DataFrame:
    grouped: defaultdict[tuple[str, str], list[ReportRecord]] = defaultdict(list)
    rows: list[dict[str, object]] = []
    for record in report_records:
        grouped[(record.model_name, "Total")].append(record)
        grouped[(record.model_name, record.modality)].append(record)

    for model_name in MODEL_ORDER:
        for group_name in GROUP_ORDER:
            records = grouped.get((model_name, group_name), [])
            if not records:
                continue
            tp = sum(1 for r in records if r.pred_has_error and r.true_has_error)
            fp = sum(1 for r in records if r.pred_has_error and not r.true_has_error)
            fn = sum(1 for r in records if not r.pred_has_error and r.true_has_error)
            tn = sum(1 for r in records if not r.pred_has_error and not r.true_has_error)
            total = len(records)
            precision = safe_divide(tp, tp + fp)
            sensitivity = safe_divide(tp, tp + fn)
            specificity = safe_divide(tn, tn + fp)
            f1 = safe_divide(2 * precision * sensitivity, precision + sensitivity) if precision == precision and sensitivity == sensitivity else float("nan")
            acc = safe_divide(tp + tn, total)
            rows.append(
                {
                    "Model": model_name,
                    "Group": group_name,
                    "ReportCount": total,
                    "TP": tp,
                    "FP": fp,
                    "FN": fn,
                    "TN": tn,
                    "ACC": format_metric(acc),
                    "Sensitivity": format_metric(sensitivity),
                    "Precision": format_metric(precision),
                    "Specificity": format_metric(specificity),
                    "F1 score": format_metric(f1),
                }
            )
    return pd.DataFrame(rows)


def summarize_instance_binary(instance_records: list[InstanceRecord]) -> pd.DataFrame:
    grouped: defaultdict[tuple[str, str], list[InstanceRecord]] = defaultdict(list)
    rows: list[dict[str, object]] = []
    for record in instance_records:
        grouped[(record.model_name, "Total")].append(record)
        grouped[(record.model_name, record.modality)].append(record)

    for model_name in MODEL_ORDER:
        for group_name in GROUP_ORDER:
            records = grouped.get((model_name, group_name), [])
            if not records:
                continue
            total = len(records)
            tp = sum(1 for r in records if r.y_pred in ERROR_LABELS and r.y_true in ERROR_LABELS)
            fp = sum(1 for r in records if r.y_pred in ERROR_LABELS and r.y_true == "f")
            fn = sum(1 for r in records if r.y_pred == "f" and r.y_true in ERROR_LABELS)
            tn = sum(1 for r in records if r.y_pred == "f" and r.y_true == "f")
            precision = safe_divide(tp, tp + fp)
            sensitivity = safe_divide(tp, tp + fn)
            specificity = safe_divide(tn, tn + fp)
            f1 = safe_divide(2 * precision * sensitivity, precision + sensitivity) if precision == precision and sensitivity == sensitivity else float("nan")
            acc = safe_divide(tp + tn, total)
            rows.append(
                {
                    "Model": model_name,
                    "Group": group_name,
                    "InstanceCount": total,
                    "TP": tp,
                    "FP": fp,
                    "FN": fn,
                    "TN": tn,
                    "ACC": format_metric(acc),
                    "Sensitivity": format_metric(sensitivity),
                    "Precision": format_metric(precision),
                    "Specificity": format_metric(specificity),
                    "F1 score": format_metric(f1),
                }
            )
    return pd.DataFrame(rows)


def summarize_instance_multiclass(instance_records: list[InstanceRecord]) -> pd.DataFrame:
    grouped: defaultdict[tuple[str, str], list[InstanceRecord]] = defaultdict(list)
    rows: list[dict[str, object]] = []
    for record in instance_records:
        grouped[(record.model_name, "Total")].append(record)
        grouped[(record.model_name, record.modality)].append(record)

    for model_name in MODEL_ORDER:
        for group_name in GROUP_ORDER:
            records = grouped.get((model_name, group_name), [])
            if not records:
                continue
            total = len(records)
            y_true = [r.y_true for r in records]
            y_pred = [r.y_pred for r in records]
            true_counts = Counter(y_true)
            pred_counts = Counter(y_pred)
            confusion = Counter(zip(y_true, y_pred))
            correct = sum(1 for t, p in zip(y_true, y_pred) if t == p)

            precisions = []
            recalls = []
            specificities = []
            f1s = []
            for label in LABEL_ORDER:
                tp = confusion[(label, label)]
                fp = pred_counts[label] - tp
                fn = true_counts[label] - tp
                tn = total - tp - fp - fn
                precision = safe_divide(tp, tp + fp)
                recall = safe_divide(tp, tp + fn)
                specificity = safe_divide(tn, tn + fp)
                f1 = safe_divide(2 * precision * recall, precision + recall) if precision == precision and recall == recall else float("nan")
                if precision == precision:
                    precisions.append(precision)
                if recall == recall:
                    recalls.append(recall)
                if specificity == specificity:
                    specificities.append(specificity)
                if f1 == f1:
                    f1s.append(f1)

            rows.append(
                {
                    "Model": model_name,
                    "Group": group_name,
                    "InstanceCount": total,
                    "ACC": format_metric(safe_divide(correct, total)),
                    "Sensitivity": format_metric(sum(recalls) / len(recalls) if recalls else float("nan")),
                    "Precision": format_metric(sum(precisions) / len(precisions) if precisions else float("nan")),
                    "Specificity": format_metric(sum(specificities) / len(specificities) if specificities else float("nan")),
                    "F1 score": format_metric(sum(f1s) / len(f1s) if f1s else float("nan")),
                }
            )
    return pd.DataFrame(rows)


def plot_bar_panels(metrics_df: pd.DataFrame, output_png: Path, output_pdf: Path) -> None:
    fig, axes = plt.subplots(2, 3, figsize=(18.5, 11.5), sharey=True)
    axes = axes.flatten()
    for ax in axes[len(MODEL_ORDER):]:
        ax.axis("off")

    for ax, model_name, panel_letter in zip(axes, MODEL_ORDER, list("ABCDE")):
        plot_df = metrics_df[metrics_df["Model"] == model_name].copy()
        plot_df["Group"] = pd.Categorical(plot_df["Group"], categories=GROUP_ORDER, ordered=True)
        plot_df = plot_df.sort_values("Group")
        values = plot_df[METRIC_COLUMNS].astype(float).to_numpy()
        x = np.arange(len(GROUP_ORDER))
        width = 0.15
        for idx, metric_name in enumerate(METRIC_COLUMNS):
            bars = ax.bar(
                x + (idx - 2) * width,
                values[:, idx],
                width=width,
                label=metric_name,
                color=PLOT_METRIC_COLORS[idx],
                edgecolor="black",
                linewidth=0.4,
            )
            for bar, value in zip(bars, values[:, idx]):
                ax.text(bar.get_x() + bar.get_width() / 2, value + 0.01, f"{value:.2f}", ha="center", va="bottom", fontsize=7.5, rotation=90)
        ax.set_xticks(x)
        ax.set_xticklabels(GROUP_ORDER)
        ax.set_ylim(0, 1.12)
        ax.set_xlabel("Modality")
        ax.set_title(model_name, fontweight="bold", pad=10)
        ax.grid(axis="y", linestyle="--", alpha=0.25)
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.text(-0.12, 1.04, panel_letter, transform=ax.transAxes, fontsize=18, fontweight="bold", va="top")

    axes[0].set_ylabel("Performance Score")
    axes[3].set_ylabel("Performance Score")
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 1.01), ncol=5, frameon=False)
    fig.suptitle("Quality Control Performance on the Matched 10,000-Report Cohort", fontsize=18, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(output_png, dpi=500, bbox_inches="tight")
    fig.savefig(output_pdf, dpi=500, bbox_inches="tight")
    plt.close(fig)


def plot_radar_panels(metrics_df: pd.DataFrame, output_png: Path, output_pdf: Path) -> None:
    angles = np.linspace(0, 2 * np.pi, len(METRIC_COLUMNS), endpoint=False).tolist()
    angles += angles[:1]
    fig, axes = plt.subplots(2, 3, figsize=(17.5, 11), subplot_kw={"polar": True})
    axes = axes.flatten()
    for ax in axes[len(MODEL_ORDER):]:
        ax.axis("off")

    for ax, model_name, panel_letter in zip(axes, MODEL_ORDER, list("ABCDE")):
        plot_df = metrics_df[metrics_df["Model"] == model_name].copy()
        plot_df["Group"] = pd.Categorical(plot_df["Group"], categories=GROUP_ORDER, ordered=True)
        plot_df = plot_df.sort_values("Group")
        for group in GROUP_ORDER:
            row = plot_df[plot_df["Group"] == group].iloc[0]
            values = [float(row[col]) for col in METRIC_COLUMNS]
            values += values[:1]
            ax.plot(angles, values, linewidth=2, label=group, color=GROUP_COLORS[group])
            ax.fill(angles, values, alpha=0.06, color=GROUP_COLORS[group])
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(METRIC_COLUMNS)
        ax.set_ylim(0, 1.0)
        ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
        ax.set_yticklabels(["0.2", "0.4", "0.6", "0.8", "1.0"])
        ax.set_title(model_name, fontweight="bold", pad=16)
        ax.grid(alpha=0.25)
        ax.text(-0.12, 1.08, panel_letter, transform=ax.transAxes, fontsize=18, fontweight="bold", va="top")

    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 1.01), ncol=4, frameon=False)
    fig.suptitle("Radar Plot of Performance on the Matched 10,000-Report Cohort", fontsize=18, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(output_png, dpi=500, bbox_inches="tight")
    fig.savefig(output_pdf, dpi=500, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    configure_plot_style()
    report_records, instance_records, file_stats_df = load_matched_records()
    report_binary_df = summarize_report_binary(report_records)
    instance_binary_df = summarize_instance_binary(instance_records)
    instance_multiclass_df = summarize_instance_multiclass(instance_records)

    output_dir = ensure_unique_dir(ROOT / "五模型AI标签_匹配10000报告分析结果")
    output_dir.mkdir(parents=True, exist_ok=False)
    excel_path = output_dir / "five_llm_matched_10000_quality_control_comparison.xlsx"

    overview_df = pd.DataFrame(
        [
            {"Item": "Models", "Value": ", ".join(MODEL_ORDER)},
            {"Item": "Cohort definition", "Value": "Matched 10,000 reports from the three root Excel files: CT first 4000, MRI first 3000, X-ray 3000"},
            {"Item": "Primary metric basis", "Value": "Label-instance level (error1-3 all included)"},
            {"Item": "Modalities", "Value": "Total, X-ray, CT, MRI"},
            {"Item": "Matched reports across all models", "Value": len(report_records)},
            {"Item": "Matched label instances across all models", "Value": len(instance_records)},
            {"Item": "Source root", "Value": str(detect_ai_root())},
        ]
    )

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="overview", index=False)
        file_stats_df.to_excel(writer, sheet_name="file_stats", index=False)
        report_binary_df.to_excel(writer, sheet_name="report_binary_metrics", index=False)
        instance_binary_df.to_excel(writer, sheet_name="instance_binary_metrics", index=False)
        instance_multiclass_df.to_excel(writer, sheet_name="instance_multiclass_metrics", index=False)

    plot_bar_panels(
        instance_multiclass_df,
        output_png=output_dir / "sci_bar_chart_five_llms_matched_10000.png",
        output_pdf=output_dir / "sci_bar_chart_five_llms_matched_10000.pdf",
    )
    plot_radar_panels(
        instance_multiclass_df,
        output_png=output_dir / "sci_radar_chart_five_llms_matched_10000.png",
        output_pdf=output_dir / "sci_radar_chart_five_llms_matched_10000.pdf",
    )

    print(f"Output directory: {output_dir}")
    print(f"Workbook: {excel_path}")
    print(f"Bar chart: {output_dir / 'sci_bar_chart_five_llms_matched_10000.png'}")
    print(f"Radar chart: {output_dir / 'sci_radar_chart_five_llms_matched_10000.png'}")
    print(f"Matched report rows loaded: {len(report_records)}")
    print(f"Matched label instances loaded: {len(instance_records)}")


if __name__ == "__main__":
    main()
